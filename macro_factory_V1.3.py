from openpyxl import load_workbook
from openpyxl import Workbook
from bs4 import BeautifulSoup
import re
from datetime import date
import uuid
import requests
import json


def get_api_urls(input_urls):
    output_urls = []
    for url in input_urls:
        articleId = re.findall("([^/]+)/?$", url)[0]
        html = requests.get(url)
        soup = BeautifulSoup(html.text, "lxml")
        cafeId = soup.find("input", {"name": "clubid"}).get("value")
        new_url = f"https://apis.naver.com/cafe-web/cafe-articleapi/cafes/{cafeId}/articles/{articleId}/comments/pages/1?requestFrom=A&orderBy=asc"
        output_urls.append({"before_url": url, "new_url": new_url})
        print(new_url)

    return output_urls


def getCount(input_urls):
    urls = get_api_urls(input_urls)
    num = 1
    lists = []
    for url in urls:
        try:
            str = requests.get(url["new_url"]).text
            data = json.loads(str)
            # 필요한 데이터
            cafe_name = data["cafe"]["name"]
            readCount = data["article"]["readCount"]
            commentCount = data["article"]["commentCount"]
            nickname = data["article"]["writer"]["nick"]
            writer_id = data["article"]["writer"]["id"]
            scrapCount = data["article"]["scrapCount"]
            writer_level = data["article"]["writer"]["memberLevelName"]
            memberCount = data["cafe"]["memberCount"]
            title = data["article"]["subject"]
            number = num
            today = date.today()
            category = "카페바이럴"
            print({"조회수": readCount, "댓글수": commentCount,
                   "카페명": cafe_name, "닉네임": nickname, "횟수": num})
            lists.append([number, today, category, cafe_name, memberCount, url["before_url"],
                          title, readCount, commentCount, nickname, writer_id, writer_level, scrapCount])
            num = num + 1
        except KeyError:
            error = data["reason"]
            number = num
            lists.append([number, None, None, None, None,
                          url["before_url"], error])
            num = num + 1

    return lists


def createExcel(lists):
    wb = Workbook()
    ws = wb.active
    ws.title = "testSheet"
    ws.append(("NO", "작성일", "분류", "카페명", "카페 회원수", "콘텐츠 URL", "제목",
              "조회수", "댓글", "닉네임", "writer_id", "writer_level", "scrapCount", "비고"))
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 17
    ws.column_dimensions["C"].width = 17
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 60
    ws.column_dimensions["G"].width = 60
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 10
    ws.column_dimensions["J"].width = 25
    ws.column_dimensions["K"].width = 20
    ws.column_dimensions["L"].width = 20
    ws.column_dimensions["M"].width = 20
    ws.column_dimensions["N"].width = 20
    for list in lists:
        ws.append(list)
    filename = uuid.uuid4()
    wb.save(f"{filename}.xlsx")
    wb.close()


def start():
    wb = load_workbook("urlList.xlsx")
    ws = wb.active
    urls = []

    # cell 갯수를 모를 때
    for x in range(2, ws.max_row + 1):
        for y in range(2, ws.max_column + 1):
            if ws.cell(row=x, column=y).value != None:
                if y == 2:  # 업체명
                    #print(f"좌표({x}행,{y}열)값 : {ws.cell(row=x, column=y).value}")
                    url = ws.cell(row=x, column=y).value
                    # 문자열 끝부분 공백제거 추가
                    urls.append(url.rstrip())

    # print(urls)
    lists = getCount(urls)
    createExcel(lists)
    # print(lists)


start()
